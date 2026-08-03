"""Parse CSV / Excel pour l'import produits stock."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from typing import Any


HEADER_ALIASES: dict[str, str] = {
    "code": "code",
    "sku": "code",
    "ref": "code",
    "reference": "code",
    "nom": "nom",
    "name": "nom",
    "produit": "nom",
    "designation": "nom",
    "désignation": "nom",
    "unite": "unite",
    "unité": "unite",
    "unit": "unite",
    "uom": "unite",
    "seuil_min": "seuil_min",
    "seuil": "seuil_min",
    "seuil minimum": "seuil_min",
    "minimum": "seuil_min",
    "minimum_stock": "seuil_min",
    "alert_threshold": "seuil_min",
    "depot": "depot",
    "dépot": "depot",
    "entrepot": "depot",
    "entrepôt": "depot",
    "magasin": "depot",
    "quantite": "quantite",
    "quantité": "quantite",
    "qty": "quantite",
    "qte": "quantite",
    "stock": "quantite",
    "prix_achat": "prix_achat",
    "prix": "prix_achat",
    "cout": "prix_achat",
    "coût": "prix_achat",
    "unit_price": "prix_achat",
    "purchase_price": "prix_achat",
}

TEMPLATE_HEADERS = ["code", "nom", "unite", "seuil_min", "depot", "quantite", "prix_achat"]
TEMPLATE_SAMPLE = [
    ["RIZ-001", "Riz parfumé", "kg", "10", "Magasin principal", "50", "1500"],
    ["HUILE-01", "Huile végétale", "L", "5", "Magasin principal", "20", "2500"],
    ["", "Serviettes", "piece", "100", "", "", ""],
]


@dataclass
class ImportRow:
    line_number: int
    code: str | None
    nom: str
    unite: str
    seuil_min: float
    depot: str | None
    quantite: float | None
    prix_achat: float | None


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("_", " ").replace("-", " ")
    text = " ".join(text.split())
    return HEADER_ALIASES.get(text, HEADER_ALIASES.get(text.replace(" ", "_"), text.replace(" ", "_")))


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return None
    return float(text)


def rows_from_dicts(dicts: list[dict[str, Any]], *, start_line: int = 2) -> list[ImportRow]:
    rows: list[ImportRow] = []
    for offset, raw in enumerate(dicts):
        line_number = start_line + offset
        mapped = {normalize_header(key): value for key, value in raw.items() if key is not None}
        nom = _cell_str(mapped.get("nom"))
        if not nom:
            # Ligne vide complète → ignorer
            if not any(_cell_str(mapped.get(key)) for key in ("code", "unite", "depot", "quantite")):
                continue
            raise ValueError(f"Ligne {line_number}: le nom du produit est obligatoire.")
        unite = _cell_str(mapped.get("unite"))
        if not unite:
            raise ValueError(f"Ligne {line_number}: l'unité est obligatoire.")
        try:
            seuil = _parse_number(mapped.get("seuil_min"))
            quantite = _parse_number(mapped.get("quantite"))
            prix = _parse_number(mapped.get("prix_achat"))
        except ValueError as exc:
            raise ValueError(f"Ligne {line_number}: nombre invalide ({exc}).") from exc
        if quantite is not None and quantite < 0:
            raise ValueError(f"Ligne {line_number}: la quantité ne peut pas être négative.")
        if quantite == 0:
            quantite = None
        depot = _cell_str(mapped.get("depot")) or None
        if quantite is not None and not depot:
            raise ValueError(f"Ligne {line_number}: le dépôt est obligatoire quand une quantité est fournie.")
        code = _cell_str(mapped.get("code")) or None
        rows.append(
            ImportRow(
                line_number=line_number,
                code=code,
                nom=nom,
                unite=unite,
                seuil_min=float(seuil or 0),
                depot=depot,
                quantite=quantite,
                prix_achat=prix if prix is not None and prix >= 0 else None,
            )
        )
    return rows


def parse_csv_bytes(content: bytes) -> list[ImportRow]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("Fichier CSV sans en-têtes.")
    dicts = [dict(row) for row in reader]
    return rows_from_dicts(dicts, start_line=2)


def parse_xlsx_bytes(content: bytes) -> list[ImportRow]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Support Excel (.xlsx) indisponible sur le serveur. Utilisez un CSV.") from exc

    workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("Fichier Excel vide.") from exc
    headers = [normalize_header(cell) for cell in header_row]
    if "nom" not in headers:
        raise ValueError("Colonne obligatoire manquante: nom")
    dicts: list[dict[str, Any]] = []
    for values in rows_iter:
        if values is None or all(value is None or str(value).strip() == "" for value in values):
            continue
        mapped: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header or header.startswith("unnamed"):
                continue
            mapped[header] = values[index] if index < len(values) else None
        dicts.append(mapped)
    return rows_from_dicts(dicts, start_line=2)


def parse_import_file(filename: str | None, content: bytes) -> list[ImportRow]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx_bytes(content)
    if name.endswith(".xls"):
        raise ValueError("Format .xls non supporté. Enregistrez en .xlsx ou CSV UTF-8.")
    # Défaut CSV (y compris .csv / sans extension)
    if name.endswith(".csv") or not name or name.endswith(".txt"):
        return parse_csv_bytes(content)
    # Tentative intelligente
    if content[:2] == b"PK":
        return parse_xlsx_bytes(content)
    return parse_csv_bytes(content)


def build_template_csv() -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(TEMPLATE_HEADERS)
    writer.writerows(TEMPLATE_SAMPLE)
    return buffer.getvalue().encode("utf-8-sig")


def build_template_xlsx() -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Produits"
    sheet.append(TEMPLATE_HEADERS)
    for row in TEMPLATE_SAMPLE:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def slug_depot_code(name: str) -> str:
    base = re.sub(r"[^A-Za-z0-9]+", "-", (name or "").strip().upper()).strip("-")
    return (base or "DEPOT")[:40]
